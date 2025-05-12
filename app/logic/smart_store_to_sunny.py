import io
import os
from datetime import datetime

import msoffcrypto
import pandas as pd
from openpyxl import load_workbook


def convert_smart_store_to_sunny(input_path, output_path, execute_date):
    df = read_excel_file(input_path)
    converted_df = convert_and_rearrange_data(df, execute_date)
    return append_to_output_file(output_path, converted_df)


def read_excel_file(input_path):
    # 암호화된 Excel 파일 열기
    with open(input_path, "rb") as file:
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(password="1111")  # 🔑 비밀번호 입력

        decrypted = io.BytesIO()
        office_file.decrypt(decrypted)

    # 암호 해제된 메모리 스트림에서 pandas로 읽기
    return pd.read_excel(decrypted, sheet_name="발주발송관리", engine='openpyxl')


def convert_and_rearrange_data(df, execute_date):
    new_df = pd.DataFrame()

    month_str = execute_date.strftime('%Y%m')

    new_df['주문자'] = df['수취인명']
    new_df['스토어상품번호'] = df['상품번호']
    new_df['스토어상품'] = df['상품명']
    new_df['수량'] = df['수량']
    new_df['연락처'] = df['수취인연락처1']
    new_df['연락처2'] = df['수취인연락처2']
    new_df['주소'] = df['통합배송지']
    new_df['배송메모'] = df['배송메세지']

    new_df['월'] = month_str
    new_df['날짜'] = execute_date
    new_df['거래유형'] = '출고'

    return new_df


def append_to_output_file(output_path, new_df, sheet_name='장부(24.01~)'):
    # 기존 파일의 모든 시트 읽기
    sheets = pd.read_excel(output_path, sheet_name=None, engine='openpyxl')

    # 장부 시트 가져오기
    existing_df = sheets[sheet_name]

    # 기존 데이터 + 새 데이터 이어붙이기
    combined_df = pd.concat([existing_df, new_df], ignore_index=True)

    # 업데이트된 장부 시트로 교체
    sheets[sheet_name] = combined_df

    # 새 파일명 생성
    new_file = generate_new_filename(os.path.dirname(output_path))

    # 모든 시트를 새 파일로 저장
    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    # 날짜 포맷 적용
    wb = load_workbook(new_file)
    ws = wb[sheet_name]

    # 두 번째 열 (엑셀은 1-based index → column B = 2)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):  # 날짜 row만
        for cell in row:
            cell.number_format = 'mm"월" dd"일"'

    # header row 설정
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # 저장
    wb.save(new_file)

    print(f"새 파일 저장 완료: {new_file}")
    return new_file


def generate_new_filename(output_dir):
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"재고정리양식_신장부{timestamp_str}.xlsx"
    return os.path.join(output_dir, filename)
